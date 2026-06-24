from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "测试类型",
    "题目-拒答题",
    "语料及生成内容的主要安全风险项",
    "类别",
    "回答\n(以下内容仅供参考，大模型每次答案都会有所差异，大体一致)",
    "是否合格",
]


def export_excel(rows: list[dict], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "distillation"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(
            [
                row.get("test_type", "文本生成"),
                row.get("prompt", ""),
                row.get("risk_item", ""),
                row.get("category", ""),
                row.get("answer", ""),
                "合格" if row.get("judge_pass") else "不合格",
            ]
        )

    widths = [14, 44, 30, 36, 88, 14]
    header_fill = PatternFill("solid", fgColor="F4F7FA")
    border = Border(
        left=Side(style="thin", color="222222"),
        right=Side(style="thin", color="222222"),
        top=Side(style="thin", color="222222"),
        bottom=Side(style="thin", color="222222"),
    )

    for idx, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = header_fill
            else:
                cell.font = Font(size=11)

    for row_idx in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 54
        sheet.cell(row_idx, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.cell(row_idx, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sheet.row_dimensions[1].height = 48
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


def export_positive_excel(rows: list[dict], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "positive_non_refusal"
    sheet.append(["问题", "类别（标签）", "答案"])
    for row in rows:
        sheet.append([row.get("question", ""), row.get("category", ""), row.get("answer", "")])

    widths = [52, 20, 88]
    header_fill = PatternFill("solid", fgColor="F4F7FA")
    border = Border(
        left=Side(style="thin", color="222222"),
        right=Side(style="thin", color="222222"),
        top=Side(style="thin", color="222222"),
        bottom=Side(style="thin", color="222222"),
    )

    for idx, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = header_fill
            else:
                cell.font = Font(size=11)

    for row_idx in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 64
        sheet.cell(row_idx, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.cell(row_idx, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sheet.row_dimensions[1].height = 42
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path
